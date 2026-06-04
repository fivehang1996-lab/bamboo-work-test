import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('D:/claude/lsxy_data.json', 'r', encoding='utf-8') as f:
    j = json.load(f)

header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
info_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

wb = openpyxl.Workbook()
ws = wb.active

meta = j['meta']
ws.merge_cells("A1:F1")
ws.cell(row=1, column=1, value=meta['title']).font = Font(name="Microsoft YaHei", bold=True, size=14)
ws.merge_cells("A2:F2")
ws.cell(row=2, column=1, value=meta['subtitle']).font = Font(name="Microsoft YaHei", size=8, color="666666")
ws.merge_cells("A3:F3")
ws.cell(row=3, column=1, value=meta['note']).font = Font(name="Microsoft YaHei", size=9, color="006600", bold=True)

sheet_data = j['sheets']['蓝色星原-恒序测试问卷']
headers = sheet_data['headers']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = wrap; c.border = border

for i, item in enumerate(sheet_data['data'], 6):
    row = [item['seq'], item['q'], item['t'], item['o'], item['f'], item['n']]
    for col, val in enumerate(row, 1):
        c = ws.cell(row=i, column=col, value=val)
        c.alignment = wrap; c.border = border; c.font = Font(name="Microsoft YaHei", size=10)
        if "---" in str(item['t']):
            c.fill = section_fill
            c.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 60
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 52
ws.column_dimensions["E"].width = 26
ws.column_dimensions["F"].width = 30
ws.freeze_panes = "A6"
ws.title = "蓝色星原-恒序测试问卷"

# Sheet 2: comparison
ws2 = wb.create_sheet("三款游戏问卷对比")
ws2.merge_cells("A1:F1")
ws2.cell(row=1, column=1, value="蓝色星原 vs 异环 vs 卡厄思梦境 测试招募问卷结构对比").font = Font(name="Microsoft YaHei", bold=True, size=13)

for col, h in enumerate(["对比维度", "蓝色星原 恒序测试", "异环 共存测试", "卡厄思梦境 深潜者测试", "差异分析", "启示"], 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = wrap; c.border = border

compare = [
    ["测试类型", "限量不计费删档", "限量计费删档", "限量计费删档", "蓝色星原非付费；异环/卡厄思付费", "付费测需额外考察消费力分层"],
    ["问卷渠道", "多平台(B站/TapTap等)+去重初筛", "官网专属唯一链接", "B站biligame系统", "蓝色星原最开放/异环最封闭/卡厄思平台绑定", "传播策略不同反映不同筛选逻辑"],
    ["设备筛选", "PC(i7+GTX1060+16GB)+安卓(骁龙8Gen1+8GB) 门槛最高", "手机型号+RAM / PC配置", "PC(i5+GTX960+8GB)+移动端(6GB) 门槛最低", "蓝原门槛最高(150GB存储)；卡厄思最低", "设备门槛=游戏品质定位信号"],
    ["设备题细节", "CPU/GPU/RAM/分辨率均需填，视频教dxdiag方法", "手机型号+RAM，较简单", "配置下拉选择，较简要", "蓝原设备题最专业(教用户查配置)", "硬核向问卷会教用户获取设备信息"],
    ["游戏频率消费", "日均时长+周频率+月消费(4题) 最细", "周频率+月消费(2题)", "周频率+月消费(2题，推断)", "蓝原多了日均时长维度", "更细的时间分层便于筛选"],
    ["手游经历", "限选5项+逐竞品问游玩时长深度 最精准", "14种类型+22款竞品(不限制) 覆盖最广", "锚定肉鸽/卡牌+第七史诗 最聚焦", "三种策略：精准/广撒网/品类聚焦", ""],
    ["PC/主机经历", "有独立PC/主机板块(与手游分开)", "无PC游戏题", "不确定", "蓝原是双端游戏需要分别考察", ""],
    ["二次元浓度", "无ACGN题(推断)", "ACGN行为+身份自评(4题)", "不确定", "异环二次元浓度考察最深", "蓝原更重游戏经验而非文化属性"],
    ["PV/角色偏好", "无(推断)", "11角色/时装投票", "不确定", "异环独有", ""],
    ["个人信息", "手机+邮箱+实名+性别年龄职业(4题)", "手机+邮箱+PWG+实名+协议(5题)", "手机+QQ/微信+邮箱+实名(4题)", "三者相似，行业标准模板", ""],
    ["摘录完整度", "4.5/5星 两份攻略视频互补,OCR质量中上", "5/5星 完整录屏+高质量OCR", "2/5星 视频非问卷录屏", "蓝原接近异环摘录质量", "两份视频互补是关键"],
]

for i, row_data in enumerate(compare, 4):
    for col, val in enumerate(row_data, 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.alignment = wrap; c.border = border; c.font = Font(name="Microsoft YaHei", size=10)

for i, w in enumerate([14, 30, 28, 28, 32, 30], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

out = "D:/claude/蓝色星原_恒序测试_问卷摘录.xlsx"
wb.save(out)
print(f"OK -> {out}")
