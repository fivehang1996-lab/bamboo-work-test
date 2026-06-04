import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
wrap_align = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
neutral_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: 跨组情感矩阵
# ============================================================
ws1 = wb.active
ws1.title = "跨组情感矩阵"

# Dimension scores: [ARPG, 银恶, 归龙潮, DNF火影, 苍翼]
# -2=强烈负面, -1=负面, 0=中性, +1=正面, +2=强烈正面
dimensions = {
    "操作流畅性":    [-2, -2, -2, -2, -1],  # 所有组都差，苍翼稍好
    "追飞船/无人机": [-2, -2, -2, -2, -1],  # 所有组都提到，核心痛点
    "剧情吸引力":    [-2, -2, -2, -1, -1],  # ARPG/银恶/归龙潮强烈负面
    "角色切换/策略": [-2, -1, -1, -1, -1],  # ARPG组最差
    "指引/教学":     [-2, -1, -2, -2, -1],  # 归龙潮和DNF火影最差
    "战斗打击感":    [-1,  0, -1, -1,  0],  # 银恶和苍翼OK
    "战斗乐趣":      [ 0, +1,  0,  0, +1],  # 银恶和苍翼正面
    "角色满意度":    [ 1, +1,  0,  0, +2],  # 苍翼最高
    "场景美术":      [-1,  0, -1, -2, -1],  # DNF火影最差
    "探索体验":      [ 0, +2, -1, -1,  0],  # 银恶组极高评价
    "目标感":        [-1, +1, -1, -1,  0],  # 银恶组最好
    "2D+二次元新颖": [+1, +1,  0, +1, +1],  # 全组基本正面
    "整体满意度":    [-1,  0, -1, -1,  0],  # 整体偏负面
    "上手难度":      [ 0,  0,  0,  0,  0],  # 全组中性
    "视角/镜头":     [-1,  0, -1, -1, -1],  # 多数负面
}

groups = ["ARPG组", "银恶组", "归龙潮组", "DNF火影组", "苍翼组"]
group_desc = [
    "原神/二次元开放世界玩家",
    "类银河恶魔城硬核玩家",
    "归龙潮流失玩家",
    "DNF/火影格斗玩家",
    "苍翼/横版动作玩家",
]

# Title
ws1.merge_cells("A1:H1")
ws1.cell(row=1, column=1, value="《绿梦》首日核心CE — 五组跨组情感矩阵").font = Font(name="微软雅黑", bold=True, size=14)

# Headers
headers = ["评估维度", "ARPG组", "银恶组", "归龙潮组", "DNF火影组", "苍翼组", "跨组共识度", "综合评级"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = wrap_align
    c.border = thin_border

# Sub-header with descriptions
for col, desc in enumerate(group_desc, 2):
    c = ws1.cell(row=4, column=col, value=desc)
    c.font = Font(name="微软雅黑", size=8, italic=True, color="666666")
    c.alignment = wrap_align
    c.border = thin_border

def consensus_score(scores):
    """Calculate consensus: how many groups agree on direction"""
    signs = []
    for s in scores:
        if s > 0: signs.append(1)
        elif s < 0: signs.append(-1)
        else: signs.append(0)
    # Count the majority sign
    pos = sum(1 for s in signs if s > 0)
    neg = sum(1 for s in signs if s < 0)
    neu = sum(1 for s in signs if s == 0)
    majority = max(pos, neg, neu)
    return majority  # 3-5

def overall_rating(scores):
    avg = sum(scores) / len(scores)
    if avg <= -1.5: return "🔴 严重问题"
    elif avg <= -0.5: return "🟠 需改进"
    elif avg < 0.5: return "🟡 中性/分歧"
    elif avg < 1.5: return "🟢 亮点"
    else: return "🟢🟢 核心优势"

for i, (dim, scores) in enumerate(dimensions.items(), 5):
    row = [dim]
    for s in scores:
        if s == -2: row.append("强烈负面(-2)")
        elif s == -1: row.append("负面(-1)")
        elif s == 0: row.append("中性(0)")
        elif s == 1: row.append("正面(+1)")
        elif s == 2: row.append("强烈正面(+2)")

    cons = consensus_score(scores)
    if cons >= 4:
        row.append(f"高度一致({cons}/5)")
    elif cons >= 3:
        row.append(f"基本一致({cons}/5)")
    else:
        row.append(f"有分歧({cons}/5)")

    row.append(overall_rating(scores))

    for col, val in enumerate(row, 1):
        c = ws1.cell(row=i, column=col, value=val)
        c.alignment = wrap_align
        c.border = thin_border
        c.font = Font(name="微软雅黑", size=10)
        # Color the score cells
        if "强烈负面" in val:
            c.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        elif "负面" in val:
            c.fill = red_fill
        elif "强烈正面" in val:
            c.fill = PatternFill(start_color="57BB8A", end_color="57BB8A", fill_type="solid")
        elif "正面" in val:
            c.fill = green_fill
        elif "高度一致" in val:
            c.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            c.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        elif "基本一致" in val:
            c.fill = yellow_fill

ws1.column_dimensions["A"].width = 18
for i in range(2, 7):
    ws1.column_dimensions[get_column_letter(i)].width = 16
ws1.column_dimensions["G"].width = 18
ws1.column_dimensions["H"].width = 18

# ============================================================
# Sheet 2: 跨组痛点Ranking
# ============================================================
ws2 = wb.create_sheet("痛点Ranking")

ws2.merge_cells("A1:F1")
ws2.cell(row=1, column=1, value="跨组痛点严重度排名（按提及组数×严重度排序）").font = Font(name="微软雅黑", bold=True, size=13)

headers2 = ["排名", "痛点", "提及组数(5组)", "综合严重度", "各组表现", "典型引文"]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = wrap_align
    c.border = thin_border

pain_points = [
    (1, "追飞船/无人机难度过高", "5/5", "⭐⭐⭐⭐⭐ 核心痛点",
     "ARPG:死30遍崩溃 | 银恶:讲一次喷一次 | 归龙潮:难度直线上升 | DNF:乐趣断崖 | 苍翼:追20分钟",
     "黄炳桥:追了30遍飞船之后没有给我宝箱，没有给我正反馈，我特别崩溃，我甚至想我能不能不玩这个游戏了"),
    (2, "操作延迟/不跟手/卡顿", "5/5", "⭐⭐⭐⭐⭐ 核心痛点",
     "ARPG:0.2秒延迟+二段跳X轴缺 | 银恶:3-4帧延迟 | 归龙潮:冲刺反方向Bug | DNF:方向键回位Bug | 苍翼:按键丢失",
     "黄炳桥:永远有0.2秒左右的延迟，这是无法接受的"),
    (3, "剧情薄弱/角色扁平", "5/5", "⭐⭐⭐⭐⭐ 核心痛点",
     "ARPG:九成废话OOC | 银恶:角色是死的 | 归龙潮:千篇一律 | DNF:无CG | 苍翼:剧情接不上",
     "孙鑫:主线剧情九成以上废话……中间全是废话，不如把营地拉短一点，中间全删了吧"),
    (4, "指引/教学缺失", "5/5", "⭐⭐⭐⭐ 强烈",
     "ARPG:太差了 | 银恶:技能机制不教 | 归龙潮:一塌糊涂 | DNF:传送/升级未教 | 苍翼:术语无解释",
     "何美聪:太差了"),
    (5, "角色切换无策略感", "5/5", "⭐⭐⭐⭐ 强烈",
     "ARPG:仅血包和冷却 | 银恶:亮了就点 | 归龙潮:躲避/亮就按 | DNF:不知目的 | 苍翼:设计意图有但未教",
     "黄炳桥:切角色的唯一目的就是因为这个角色没有血了"),
    (6, "角色/怪太小，场景无重点", "4/5", "⭐⭐⭐",
     "ARPG:护盾看不清楚 | 银恶:角色占比大 | 归龙潮:未提及 | DNF:视角不能切 | 苍翼:通关时太小",
     "孙鑫:无论玩家角色还是怪都有点小……你不仔细盯着看，看不到它"),
    (7, "养成路径不透明", "4/5", "⭐⭐⭐",
     "ARPG:开宝箱精度为零 | 银恶:略提 | 归龙潮:材料不知用途 | DNF:1级打20级Boss | 苍翼:未提及",
     "黄炳桥:我开了N个宝箱获取精度是零，我不知道哪里获取突破材料……我过了一个小时才一级"),
    (8, "画面粗糙/塑料感", "4/5", "⭐⭐⭐",
     "ARPG:略提 | 银恶:粗糙 | 归龙潮:画面精细度太低 | DNF:不像2025年游戏 | 苍翼:贴图感/模糊",
     "金申杰:画风不是2025年做出来的游戏，10年前我会觉得这个游戏还不错"),
    (9, "二段跳手感差(X轴/卡平台)", "4/5", "⭐⭐⭐",
     "ARPG:无X轴偏移 | 银恶:未特别提及 | 归龙潮:跳跃延迟 | DNF:三段跳建议 | 苍翼:不明显",
     "黄炳桥:二段跳没有进行X轴上的偏移，第一段跳可以往指定方向跳，二段跳只能原地向上跳"),
    (10, "打击感弱/无振动反馈", "3/5", "⭐⭐",
     "ARPG:缺振动 | 银恶:OK | 归龙潮:音效好 | DNF:弱 | 苍翼:不如苍翼软",
     "陆少杰:打击感最差，整体来说打击感的振动反馈没有"),
]

for i, (rank, point, groups_mentioned, severity, group_detail, quote) in enumerate(pain_points, 4):
    row = [rank, point, groups_mentioned, severity, group_detail, quote]
    for col, val in enumerate(row, 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.alignment = wrap_align
        c.border = thin_border
        c.font = Font(name="微软雅黑", size=10)
        if rank <= 3:
            c.fill = red_fill

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 26
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 55
ws2.column_dimensions["F"].width = 55

# ============================================================
# Sheet 3: 组间差异洞察
# ============================================================
ws3 = wb.create_sheet("组间差异洞察")
ws3.merge_cells("A1:G1")
ws3.cell(row=1, column=1, value="五组玩家画像差异及对评价的影响").font = Font(name="微软雅黑", bold=True, size=13)

headers3 = ["组别", "玩家特征", "核心参照系", "最敏感维度", "最宽容维度", "独有反馈", "对产品的启示"]
for col, h in enumerate(headers3, 1):
    c = ws3.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = wrap_align
    c.border = thin_border

insights = [
    ("ARPG组", "原神/鸣潮二次元玩家，佛系+收藏向", "原神/鸣潮",
     "剧情质量/角色辨识度",
     "操作韧性（死了还想挑战）",
     "关注养成和收集系统（角色+皮肤），对二段跳X轴有专业分析",
     "需要做好角色辨识度和剧情，否则留不住二次元核心用户"),
    ("银恶组", "类银河恶魔城硬核玩家，空洞/死亡细胞/奥日经验", "空洞骑士/死亡细胞/奥日",
     "地图设计/探索深度",
     "整体战斗设计（给高分因为看出设计师用心）",
     "对隐藏墙/地图设计会心一笑，深入讨论完美闪避奖励和切人策略",
     "探索和地图设计是最大加分项，设计足够用心可获硬核玩家认可"),
    ("归龙潮组", "归龙潮流失玩家，多平台杂食", "归龙潮",
     "画面品质/操作流畅",
     "无（整体偏负面）",
     "对归龙潮有全面对比框架，关注数值膨胀/付费模式/更新频率",
     "这群玩家已是流失状态，对新游戏容忍度低，但对比框架清晰"),
    ("DNF火影组", "格斗/IP驱动玩家，火影/DNF经验", "DNF/火影忍者",
     "画面时代感/装备掉落",
     "跑酷+格斗的新颖组合",
     "希望有装备掉落可视化和PK模式，对养成和每日任务有明确预期",
     "需要清晰的养成和装备获取路径，否则缺乏长线目标感"),
    ("苍翼组", "苍翼/横版动作核心玩家，3D眩晕者", "苍翼",
     "打击感/特效品质",
     "角色XP和立绘",
     "发现破盾→合轴的设计意图，主动理解游戏策略深度",
     "战术深度党，给他足够的战斗系统深度能自己挖掘，但需要巧妙提示"),
]

for i, row_data in enumerate(insights, 4):
    for col, val in enumerate(row_data, 1):
        c = ws3.cell(row=i, column=col, value=val)
        c.alignment = wrap_align
        c.border = thin_border
        c.font = Font(name="微软雅黑", size=10)

for i, w in enumerate([10, 28, 22, 18, 18, 40, 40], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 4: 优先级行动建议
# ============================================================
ws4 = wb.create_sheet("行动建议")
ws4.merge_cells("A1:E1")
ws4.cell(row=1, column=1, value="基于五组CE的行动优先级建议").font = Font(name="微软雅黑", bold=True, size=13)

headers4 = ["优先级", "问题", "不改的后果", "建议方案", "验证方式"]
for col, h in enumerate(headers4, 1):
    c = ws4.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = wrap_align
    c.border = thin_border

actions = [
    ("P0 致命", "追飞船/无人机难度",
     "5组全部在此流失/崩溃，首日留存直接崩盘",
     "① 降速+减少转弯; ② 中途加存档点; ③ 完成后给宝箱正反馈; ④ 可选跳过",
     "下一轮CE验证通过率"),
    ("P0 致命", "操作延迟/不跟手",
     "所有组的操作评分垫底，直接影响战斗和跑图核心体验",
     "① 优先修帧率(默认60帧); ② 修跳跃延迟; ③ 修二段跳X轴; ④ 修方向键复位Bug",
     "帧率测试+延迟量化+下轮CE"),
    ("P0 致命", "剧情/角色塑造",
     "五组一致负面，无人愿意分享剧情故事",
     "① 精简营地对话(删50%); ② 增加一场类似'吃饭'的生活戏; ③ 立女主第一印象; ④ 换配音",
     "下轮CE剧情跳过率对比"),
    ("P1 重要", "指引/教学系统",
     "玩家不知道怎么升级/换技能/切人策略，体验不完整",
     "① 非强制式教学(苍翼组建议巧妙融入); ② 解释共振/噪波; ③ 教传送; ④ 教升级路径",
     "下轮CE完成率对比"),
    ("P1 重要", "角色切换策略感",
     "全部组都认为切人=血包/冷却，浪费了已设计的合轴/破盾系统",
     "① 新手关强制引导破盾+合轴循环; ② 增加角色组合技/羁绊Buff; ③ 增加切人出场技",
     "下轮CE切人频率和策略感评分"),
    ("P2 改善", "场景/背景辨识度",
     "背景与可交互物混淆、地刺当草地、画面粗糙",
     "① 提高陷阱与背景的对比度; ② 加高亮提示可攀爬/可破坏物; ③ 旋转楼梯品质做标杆",
     "下轮CE场景辨识测试"),
    ("P2 改善", "养成/材料可视化",
     "开宝箱不知用途，1小时1级，缺少成长反馈",
     "① 宝箱获得时弹窗说明用途; ② 早期强制升1级; ③ 怪物掉材料可视化",
     "下轮CE养成完成率"),
    ("P3 锦上添花", "打击感/振动反馈",
     "部分玩家(尤其火影/鸣潮玩家)期望横版格斗有振动",
     "① 加入振动反馈开关; ② 优化受击音效层次",
     "下轮CE打击感评分"),
    ("P3 锦上添花", "探索/跑图比例",
     "苍翼组和银恶组认为跑图>战斗，希望多打怪",
     "① 调整跑图:战斗比例; ② 增加可选战斗遭遇",
     "下轮CE比例满意度"),
    ("P3 锦上添花", "性别多样性",
     "两处提及只有女角色→受众受限",
     "评估添加男角色的可行性和排期",
     "市场调研/目标受众验证"),
]

for i, (priority, problem, consequence, solution, verification) in enumerate(actions, 4):
    row = [priority, problem, consequence, solution, verification]
    for col, val in enumerate(row, 1):
        c = ws4.cell(row=i, column=col, value=val)
        c.alignment = wrap_align
        c.border = thin_border
        c.font = Font(name="微软雅黑", size=10)
        if "P0" in priority:
            c.fill = red_fill

for i, w in enumerate([10, 22, 40, 44, 30], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# -- Save --
outpath = "D:/claude/跨组对比矩阵.xlsx"
wb.save(outpath)
print(f"OK -> {outpath}")
