import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

hfont = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
sfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "卡厄思梦境-深潜者测试问卷"

ws.merge_cells("A1:F1")
ws.cell(row=1, column=1, value="《卡厄思梦境》「深潜者」测试招募问卷 — 已知题目摘录").font = Font(name="微软雅黑", bold=True, size=14)
ws.merge_cells("A2:F2")
ws.cell(row=2, column=1, value="来源：B站视频 BV1u8FpzwEQ9 + BV1DWFnz4ENo OCR + 官网公告(4399/B站动态/奇游) + 网络搜索 | 限量计费删档 | PC+移动端 | 仅18周岁+ | 招募：2026.2.6-3.3 | 开测：2026.3.11").font = Font(name="微软雅黑", size=8, color="666666")
ws.merge_cells("A3:F3")
ws.cell(row=3, column=1, value="注意：两个B站视频均为资讯解说+充值返还+个人体验分享，未逐题录制问卷填写过程。以下为视频OCR片段+多来源拼凑，标注了来源可信度。不如异环那份完整。").font = Font(name="微软雅黑", size=9, color="CC0000", bold=True)

headers = ["序号", "题目/内容", "题型", "选项", "来源/可信度", "备注"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=col, value=h)
    c.font = hfont; c.fill = hfill; c.alignment = wrap; c.border = border

data = []

def S(seq, title):
    data.append((seq, title, "--- 问卷板块 ---", "", "综合", ""))

def Q(seq, question, qtype, options, source, note=""):
    data.append((seq, question, qtype, options, source, note))

S("A", "问卷声明（首页）")

Q("A1",
  "【问卷填写声明】\n您好，非常感谢您参与《卡厄思梦境》[深潜者]测试的官方问卷招募活动。\n本次测试为限量计费删档测试，测试期间将会开放充值功能。\n正式公测时将会根据您在本次测试中的付费情况进行充值返还。\n仅面向18周岁及以上玩家开放。请如实填写个人信息，提交后不可修改。",
  "声明",
  "同意 / 不同意（不同意则无法继续）",
  "官方公告+视频",
  "视频展示了招募开始时间和声明页面")

S("B", "第一部分：设备信息（筛选适配）")

Q("B1",
  "请问您在本次测试中计划使用的设备是？",
  "单选题",
  "A. 主要使用移动设备（手机/平板）\nB. 主要使用PC设备\nC. 均会使用，移动设备用的多一些\nD. 均会使用，PC设备用的多一些\nE. 目前还不确定",
  "视频OCR片段+同类模板",
  "视频f01080-f01296帧OCR到设备选择页。重要：iOS用户本次无法充值，需用安卓/PC端充值")

Q("B2",
  "PC端配置要求（选择PC时出现）：\n- 最低：Intel i5 + GTX960 + 8GB RAM + Win10 64bit + 6GB存储\n- 推荐：Intel i7 + GTX1660 + 16GB RAM + Win10 64bit + 6GB存储",
  "填空/下拉",
  "CPU型号 / 显卡型号 / 运行内存 / 操作系统 / 存储空间",
  "4399+奇游 官方公告",
  "来源：act-web.biligame.com问卷页 + 4399论坛51986324")

Q("B3",
  "移动端配置要求（选择移动设备时出现）：\n- Android最低：Android 7.0 + 6GB RAM\n- Android推荐：Android 10.0 + 8GB RAM\n- iOS：待官方公布",
  "填空/下拉",
  "手机型号 / 系统版本 / 运行内存(RAM)",
  "奇游+52pk 官方公告",
  "")

S("C", "第二部分：游戏经历（筛选目标用户）")

Q("C1",
  "请问您最近半年，每周使用手机/平板/PC进行游戏的频率大概是？",
  "单选题",
  "每天都玩\n每周玩5-6天\n每周玩3-4天\n每周玩1-2天\n偶尔玩，每周不超过一天\n基本不用手机/平板玩游戏",
  "视频OCR片段+同类模板",
  "视频f01080帧OCR到相关文字")

Q("C2",
  "请问最近一年，您平均每个月在游戏方面的支出有多少？",
  "单选题",
  "0元或大部分时候不会充值\n1~100元\n101~200元\n201~500元\n501~1000元\n1001~5000元\n5000元及以上",
  "视频OCR片段+标准模板",
  "")

Q("C3",
  "【多选题】请问以下游戏类型，哪些是您经常玩的？（重点考察Roguelike/肉鸽卡牌经验）",
  "多选题",
  "Roguelike/肉鸽卡牌（杀戮尖塔、怪物火车、欺诈之地等）\n二次元角色扮演（原神、崩坏星穹铁道、绝区零等）\n卡牌策略（炉石传说、影之诗、游戏王等）\n回合制RPG（第七史诗、FGO等）\n动作RPG（崩坏3、战双帕弥什等）\nMOBA（王者荣耀、英雄联盟手游等）\n塔防（明日方舟、王国保卫战等）\n射击竞技（和平精英、三角洲行动等）\nMMORPG\n放置类\n其他",
  "网络搜索+开发商背景推断",
  "开发商为Smilegate(第七史诗团队)，卡厄思梦境为Roguelike卡牌RPG。问卷大概率重点考察肉鸽卡牌经验")

Q("C4",
  "【多选题】请问以下游戏，哪些是您深度体验过或仍然在玩的？（重点锚定竞品）",
  "多选题",
  "第七史诗（同开发商Smilegate出品）\n杀戮尖塔（Roguelike卡牌标杆）\n明日方舟\n原神\n崩坏：星穹铁道\n绝区零\n鸣潮\n幻塔\n怪物火车\n其他Roguelike卡牌，请填写：____",
  "网络搜索+合理推断",
  "第七史诗是同一开发商的成功作品，问卷极可能用它作为参照")

Q("C5",
  "请问您是否有玩过Roguelike/肉鸽类游戏的经验？如有，请描述您的游玩时长和最常玩的游戏。",
  "填空/简答",
  "开放式回答",
  "网络搜索(奇游/豌豆荚攻略)",
  "社区攻略强调：如实填写肉鸽卡牌经验可提高中选概率")

S("D", "第三部分：个人信息")

Q("D1",
  "请填写您的手机号码（资格通知渠道之一）",
  "填空",
  "11位手机号",
  "官方公告",
  "资格通过站内信+短信双通道发放")

Q("D2",
  "请填写您的QQ或微信账号（用于登录游戏）",
  "填空",
  "QQ号 / 微信号",
  "官方公告",
  "资格绑定账号，不可转让/共享/售卖")

Q("D3",
  "请填写您的联系邮箱（资格通知凭证之一）",
  "填空",
  "邮箱地址",
  "官方公告",
  "")

Q("D4",
  "实名认证信息",
  "填空",
  "真实姓名 / 身份证号（需与后续认证一致）",
  "官方公告",
  "仅限18周岁以上")

S("E", "补充信息（从视频OCR和官方动态提取）")

Q("E1",
  "招募基本信息：\n- 招募开始：2026年2月6日 10:00 (UTC+8)\n- 招募截止：2026年3月3日 23:59 (UTC+8)\n- 测试开启：2026年3月11日\n- 测试类型：限量计费删档测试\n- 测试平台：移动端（安卓+iOS）+ PC端(Windows)，双端数据互通",
  "信息",
  "—",
  "官方B站动态+视频",
  "视频中多次确认这些信息")

Q("E2",
  "充值返还规则（视频详细展示）：\n累计充值<=1500元部分 -> 返还200%免费水晶\n1500<累计<=3000元部分 -> 返还150%免费水晶\n累计>3000元部分 -> 返还125%免费水晶\n兑换比例：1元人民币 = 10免费水晶\n\niOS用户本次测试无法充值（苹果系统限制），需用安卓或PC端登录充值",
  "信息",
  "—",
  "视频f01728-f02916详细展示",
  "两个视频都花了大量篇幅展示充值返还规则")

Q("E3",
  "死灵之书（月卡）返还规则：\n1. 购买金额计入累计充值，按上述比例返还免费水晶\n2. 返还[死灵之书兑换券]，公测后可兑换30天权益（不包含付费水晶*300），上限2次",
  "信息",
  "—",
  "视频f02340-f02520详细展示",
  "")

Q("E4",
  "阿基亚农补给（通行证）返还规则：\n1. 购买金额计入累计充值\n2. 按购买版本返还[特殊资料兑换券]或[零之资料兑换券]，上限1次",
  "信息",
  "—",
  "视频f02610-f02700详细展示",
  "")

Q("E5",
  "资格通知方式：招募结束后通过站内信+短信通知获得资格的领航者",
  "信息",
  "—",
  "视频f01620帧OCR",
  "资格仅限于本人使用，严禁转让/共享/售卖/借号")

Q("E6",
  "问卷链接：https://act-web.biligame.com/act-server/wj-h5/KP_QUE_Z_2026020326611（需B站账号登录后填写）",
  "信息",
  "—",
  "视频f01170帧直接OCR到URL",
  "")

for i, (seq, question, qtype, options, source, note) in enumerate(data, 6):
    row = [seq, question, qtype, options, source, note]
    for col, val in enumerate(row, 1):
        c = ws.cell(row=i, column=col, value=val)
        c.alignment = wrap; c.border = border; c.font = Font(name="微软雅黑", size=10)
        if "---" in str(qtype):
            c.fill = sfill
            c.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        elif "声明" in str(qtype):
            c.fill = warn_fill

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 60
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 48
ws.column_dimensions["E"].width = 26
ws.column_dimensions["F"].width = 26
ws.freeze_panes = "A6"

# Sheet 2: 对比异环
ws2 = wb.create_sheet("与异环问卷对比")
ws2.merge_cells("A1:E1")
ws2.cell(row=1, column=1, value="卡厄思梦境 vs 异环 付费测试招募问卷结构对比").font = Font(name="微软雅黑", bold=True, size=13)

for col, h in enumerate(["对比维度", "异环 共存测试", "卡厄思梦境 深潜者测试", "差异点", "启示"], 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = hfont; c.fill = hfill; c.alignment = wrap; c.border = border

compare = [
    ("问卷渠道", "官网专属唯一链接", "B站biligame问卷系统", "异环独立站 / 卡厄思走B站渠道", "B站渠道触达二次元核心用户"),
    ("设备筛选", "手机型号+RAM / PC详细配置", "手机+PC配置（较简要）", "异环设备题更细", "相似结构，卡厄思移动端门槛更低"),
    ("游戏经历", "14种游戏类型+22款竞品列表", "重点锚定肉鸽/卡牌+第七史诗", "异环广撒网 / 卡厄思精准锚定1-2品类", "卡厄思更聚焦品类匹配"),
    ("二次元浓度", "ACGN行为+身份自评(4题)", "不确定是否有类似题目", "异环有完整二次元浓度测试", "卡厄思问卷可能更短更聚焦"),
    ("游戏价值偏好", "11个维度限选5项", "未知（可能侧重肉鸽/策略偏好）", "异环有详细偏好测试", "卡厄思可能没有或不相同"),
    ("PV/角色偏好", "11个新角色/时装审美投票", "不确定是否有", "异环有创意素材测试", "—"),
    ("线下/社媒偏好", "线下活动+社媒内容偏好", "不确定是否有", "异环有运营策略题", "—"),
    ("充值返还", "问卷中不明显", "视频大量展示(200%/150%/125%层级)", "卡厄思返利结构更复杂", "计费测试需清晰传达返利规则"),
    ("保密要求", "否", "否", "两者均非保密测试", "与王者荣耀世界形成对比"),
    ("摘录完整度", "5/5星 完整逐题提取", "2/5星 拼凑（视频非问卷录屏）", "异环有完整录屏视频 / 卡厄思视频是资讯解说", "方法论的局限性：必须有问卷填写录屏视频才能逐题摘录"),
]

for i, (dim, yh, ke, diff, insight) in enumerate(compare, 4):
    for col, val in enumerate([dim, yh, ke, diff, insight], 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.alignment = wrap; c.border = border; c.font = Font(name="微软雅黑", size=10)

for i, w in enumerate([14, 30, 30, 32, 30], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

out = "D:/claude/卡厄思梦境_深潜者测试_问卷摘录.xlsx"
wb.save(out)
print(f"OK -> {out}")
