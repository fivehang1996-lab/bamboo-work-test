import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
q_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
opt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
wrap_align = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "异环共存测试招募问卷"

# -- Title --
ws.merge_cells("A1:E1")
ws.cell(row=1, column=1, value="《异环》「共存测试」招募问卷 — 完整题目摘录").font = Font(name="微软雅黑", bold=True, size=14)
ws.merge_cells("A2:E2")
ws.cell(row=2, column=1, value="来源：B站视频 BV1zoB1BxE8J (PhilJia_0513) + 官方公告补充 | 测试类型：限量删档计费测试 | 平台：PC/移动端数据互通 | 仅限18周岁以上").font = Font(name="微软雅黑", size=9, color="666666")

headers = ["序号", "题目/内容", "题型", "选项", "备注/来源"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = wrap_align
    c.border = thin_border

data = []

def add_section(seq, title):
    data.append((seq, title, "--- 问卷板块 ---", "", ""))

def add_q(seq, question, qtype, options, note=""):
    data.append((seq, question, qtype, options, note))

# === 问卷声明 ===
add_section("A", "问卷填写声明（首页）")

add_q("A1", "【问卷填写声明】\n您好，非常感谢您参与《异环》[共存测试]的官方问卷招募活动。\n本次测试仅招募【18周岁及以上】玩家。填写问卷将有机会获得此次测试资格。\n请根据您实际情况认真如实填写，确保问卷信息的真实准确。\n感谢您的配合。该等测试资格发放由项目组具体筛选确认，\n您理解并同意，填写问卷并不必然获取测试资格。\n本次测试为PC端/移动端限量删档计费测试，PC端与移动端数据互通。\n我们将基于筛选测试资格等目的收集问卷信息，请您放心，\n相关信息我们将妥善保存并严格保密，以确保您的信息与数据安全。",
      "声明", "同意 / 不同意（不同意则无法继续）", "视频帧: f00135 | OCR识别")

# === 第一部分：设备信息 ===
add_section("B", "第一部分：设备信息")

add_q("B1", "请问您在本次测试中计划使用的设备是？",
      "单选题",
      "A. 主要使用移动设备（手机、平板）参与测试\nB. 主要使用PC设备参与测试\nC. 均会使用，移动设备用的多一些\nD. 均会使用，PC设备用的多一些\nE. 目前还不确定",
      "视频帧: f00360, f00585")

add_q("B2", "请问您的手机型号是？（选择移动设备时出现）",
      "下拉选择/填空",
      "例如：iPhone 13 Pro（视频中示例）\n需具体到手机型号",
      "视频帧: f00585")

add_q("B3", "请问您的手机运行内存(RAM)为？\n（查看方法：小米：设置→我的设备→运行内存；其他型号手机查看路径相似）",
      "单选题",
      "4GB及以下\n6GB\n8GB\n12GB\n16GB及以上\n不清楚",
      "视频帧: f00360, f00585")

add_q("B4", "PC设备配置信息（选择PC时出现）：",
      "填空/下拉",
      "CPU型号（如：i7-12700）\n显卡型号（如：RTX 3060）\n运行内存（如：16GB）\n操作系统版本（如：Windows 10 64位）",
      "官方公告补充：最低i7-10700/GTX1660/16GB；推荐i7-12700/RTX3060/32GB")

add_q("B5", "iOS设备信息（选择iOS时出现）：",
      "填空/下拉",
      "iPhone型号\niOS系统版本",
      "官方补充：最低iPhone12 Pro Max/iOS15")

# === 第二部分：游戏行为 ===
add_section("C", "第二部分：游戏行为与偏好")

add_q("C1", "请问您最近半年，您每周使用手机/平板进行游戏的频率大概是？",
      "单选题",
      "每天都玩\n每周玩5-6天\n每周玩3-4天\n每周玩1-2天\n偶尔玩，每周不超过一天\n基本不用手机/平板玩游戏",
      "视频帧: f00720")

add_q("C2", "请问最近一年，您平均每个月在游戏方面的支出有多少？",
      "单选题",
      "0元或大部分时候不会充值\n1~100元\n101~200元\n201~500元\n501~1000元\n1001~5000元\n5000元及以上",
      "视频帧: f00720")

add_q("C3", "【多选题】请问以下游戏类型，哪些是您经常玩的？",
      "多选题",
      "射击竞技（和平精英、三角洲行动等）\n音舞类（LoveLive、喵斯快跑、PJSK等）\n休闲益智（开心消消乐、梦幻家园、愤怒的小鸟）\n沙盒（迷你世界、我的世界等）\n二次元角色扮演（原神、崩坏：星穹铁道、绝区零）\nMOBA（王者荣耀、英雄联盟手游等）\n集换/构筑卡牌（炉石传说、宝可梦集换式卡牌、影之诗等）\n塔防（植物大战僵尸、王国保卫战等）\nMMORPG（逆水寒手游、冒险岛手游、梦幻西游等）\n体育&竞速（QQ飞车手游、巅峰极速、实况足球等）\nSLG（部落冲突、率土之滨、三国谋定天下等）\n非对称竞技（猫和老鼠、第五人格等）\n放置类（出发吧麦芬、剑与远征：启程等）\n其他，请填写：____",
      "视频帧: f00945")

add_q("C4", "【多选题】请问以下游戏，哪些是您深度体验过或仍然在玩的？",
      "多选题",
      "命运-冠位指定（FGO）\n碧蓝航线\n明日方舟\n战双帕弥什\n原神\n幻塔\n崩坏：星穹铁道\n重返未来1999\n少女前线2：追放\n无限暖暖\n鸣潮",
      "视频帧: f01215")

add_q("C5", "【多选题】请问以下游戏，哪些是您深度体验过的？",
      "多选题",
      "黑色信标\n银与绯\n星痕共鸣\n晴空之下\n胜利女神：新的希望\n火环\n辉烬\n星塔旅人\n卡拉彼丘手游\n二重螺旋\n以上都没有",
      "视频帧: f01485")

add_q("C6", "【多选题】在二次元游戏分类中，您更偏好哪些子类型？",
      "多选题",
      "二次元回合制/策略RPG（命运-冠位指定、崩坏：星穹铁道、第七史诗等）\n二次元动作ARPG（崩坏3、战双帕弥什、绝区零等）\n二次元射击（卡拉彼丘、尘白禁区、胜利女神：妮姬等）\n二次元塔防/战棋（明日方舟、火焰纹章：英雄、少女前线等）\n二次元卡牌\n二次元音舞（偶像梦幻祭、LoveLive、世界计划 缤纷舞台等）",
      "视频帧: f01755")

# === 第三部分：游戏偏好深度 ===
add_section("D", "第三部分：游戏偏好深度")

add_q("D1", "【多选题】请问您在体验游戏时，更看重哪些内容？[最多选择5项]",
      "多选题(限5项)",
      "[代入] 沉浸在游戏中，忘却现实世界的感觉\n[解谜] 享受通过思考，成功解决谜题的乐趣\n[剧情] 如剧情故事引人入胜、人物形象饱满有魅力、世界观合理且有趣\n[收集] 收集游戏中的各种道具/资产，解锁图鉴或完成成就\n[个性] 操控角色的外观/载具/房屋等拥有更多个性定制选项\n[角色] 获取喜欢的角色，体验更多与之有关的剧情\n[探索] 不断发现未知的新玩法、奖励、彩蛋\n[休闲] 能够放松和释放压力\n[挑战] 通过重复练习/研究策略才能通过的，有一定难度的各类玩法\n[画面] 如精致的画面/人物、有辨识度的美术风格\n[自由] 在剧情、地图、玩法等方面，能做更多事，能选择更多达成目标的手段",
      "视频帧: f02025")

add_q("D2", "请问近期您是否有以下行为？",
      "多选题",
      "购买动漫、游戏周边\n看日漫、国漫\n听二次元音乐（听动漫曲、VOCALOID曲）\n玩二次元游戏\n参加漫展、动漫演唱会等二次元线下活动\n关注动漫相关资讯\n参与同人创作（同人图、COS等）\n关注ACG圈子（画师圈、COS圈等）\n看MMD、MAD\n以上都没有",
      "视频帧: f02430")

add_q("D3", "请问您对于'ACGN'（动画、漫画、游戏、轻小说）的看法是？",
      "单选题",
      "ACGN是我生活的重心，我总是沉浸其中\nACGN作品深刻塑造了我的价值观或人生观\nACGN作品时常会潜移默化地影响我的观念\nACGN作品仅在某些具体事情上影响我的看法\n我很少接触ACGN，这些对我影响不大",
      "视频帧: f02430")

add_q("D4", "以下哪个描述更符合您呢？",
      "单选题（二次元身份自评）",
      "我是二次元文化的创作者，曾发布过同人作品/Cosplay等内容\n我是资深爱好者，热衷于鉴赏并向朋友'安利'优质作品\n我有感兴趣的作品，二次元是我生活娱乐的一部分，但非全部\n我会偶尔接触热门动漫作品，但我更关注三次元生活\n我对二次元文化不感兴趣，平时基本不接触相关内容",
      "视频帧: f02700")

# === 第四部分：信息获取与游戏吸引点 ===
add_section("E", "第四部分：信息获取与《异环》吸引点")

add_q("E1", "您通常通过哪些渠道获取游戏资讯？",
      "多选题",
      "游戏媒体/公众号（机核、游研社、3DM等）\n百度贴吧\n其他，请填写：____",
      "视频帧: f02970")

add_q("E2", "【多选题】您更习惯通过何种形式获取自己感兴趣的资讯？",
      "多选题",
      "官方公告\n朋友介绍\n广告曝光\n一图流\n专题文章\n资讯图文\n短视频\n直播\n社区帖子\n长视频\n其他，请填写：____",
      "视频帧: f02970")

add_q("E3", "【多选题】请问《异环》吸引您参与本次测试招募的原因是？[最多选择3项]",
      "多选题(限3项)",
      "虚幻引擎加持下的高画质与无缝加载等视觉表现\n喜欢游戏营造的都市生态，一些细节和彩蛋\n超自然元素的世界观设定\n喜欢游戏中的特定角色，请填写：____\n因为游戏里可以买车开车\n其他，请填写：____",
      "视频帧: f02970, f03240")

add_q("E4", "请问您是否看过《异环》的最新PV（《异环》「共存测试」招募PV｜「你好，新人」）？",
      "单选题",
      "看过\n没看过",
      "视频帧: f03240")

add_q("E5", "【多选题】请问以下PV中出现的新角色/角色新形象，您比较喜欢的是？",
      "多选题",
      "新角色-曳尔沐\n主角-时装1（男）\n娜娜莉-时装3\n新角色-灵可&小贞\n新角色-Exe\n新角色-伊洛伊\n娜娜莉-时装1\n娜娜莉-时装2\n主角-时装3\n新角色-卡厄斯\n新角色-真红",
      "视频帧: f03240, f03510")

# === 第五部分：线下活动与社媒偏好 ===
add_section("F", "第五部分：线下活动与社媒内容偏好")

add_q("F1", "【多选题】请问您更希望在未来看到或参与哪些形式的《异环》线下活动？",
      "多选题",
      "大型漫展（官方参展）\n周边快闪店\n餐饮主题店\n主题音乐会\n线下嘉年华（FES）\n所在城市的线下广告投放\n无人机、大型烟花等线下演出\n其他，请填写：____",
      "视频帧: f03915")

add_q("F2", "【多选题】请问您更希望在未来看到官方社媒发布哪些形式的《异环》相关内容？",
      "多选题",
      "更多形式与方面的游戏角色魅力展示\n官方或合作方分享的游戏制作幕后花絮\n通过漫画、配音小剧场等形式呈现的同一阵营下/不同阵营间的角色关系\n对于游戏世界观的额外补充说明\nPV、音乐等官方重要素材的制作花絮\n壁纸、日历等游戏主题的实用资源\n配音演员（CV）等制作人员的幕后解读或特别直播\n其他，请填写：____",
      "视频帧: f04320")

# === 第六部分：实名与联系方式（官网公告补充）===
add_section("G", "第六部分：实名信息与联系方式（官方公告补充，视频中未展示）")

add_q("G1", "请填写您的真实姓名",
      "填空",
      "需与身份证一致，仅限18周岁以上",
      "官方公告")

add_q("G2", "请填写您的手机号码",
      "填空",
      "资格通知渠道之一",
      "官方公告")

add_q("G3", "请填写您的联系邮箱",
      "填空",
      "资格通知唯一凭证，请确保准确",
      "官方公告")

add_q("G4", "完美世界游戏账号（PWG账号）",
      "填空",
      "如无可在官网注册",
      "官方公告")

add_q("G5", "【协议确认】",
      "勾选",
      "同意《完美世界游戏用户协议》\n同意《完美世界游戏个人信息保护政策》\n同意通过电话或短信方式接受通知及营销信息（可选）",
      "官方公告")

# ================================================================
# Write data rows
# ================================================================
for i, (seq, question, qtype, options, note) in enumerate(data, 5):
    row = [seq, question, qtype, options, note]
    for col, val in enumerate(row, 1):
        c = ws.cell(row=i, column=col, value=val)
        c.alignment = wrap_align
        c.border = thin_border
        c.font = Font(name="微软雅黑", size=10)
        # Color sections
        if "---" in str(qtype):
            c.fill = section_fill
            c.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        elif "声明" in str(qtype):
            c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 55
ws.column_dimensions["E"].width = 28
ws.freeze_panes = "A5"

# -- Sheet 2: Summary --
ws2 = wb.create_sheet("问卷结构概览")
ws2.merge_cells("A1:C1")
ws2.cell(row=1, column=1, value="《异环》共存测试招募问卷 — 结构概览").font = Font(name="微软雅黑", bold=True, size=13)

for col, h in enumerate(["板块", "题目数量", "核心目的"], 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = wrap_align
    c.border = thin_border

overview = [
    ("A. 问卷声明", "1题", "法律告知、年龄限制确认"),
    ("B. 设备信息", "3-5题(根据设备分支)", "筛选设备是否满足测试要求"),
    ("C. 游戏行为与偏好", "6题", "判断玩家游戏经历匹配度"),
    ("D. 游戏偏好深度", "4题", "判断二次元浓度和玩家画像"),
    ("E. 吸引点与PV感知", "5题", "判断对《异环》的兴趣深度"),
    ("F. 线下活动与社媒偏好", "2题", "运营策略参考"),
    ("G. 实名信息", "5题", "资格发放与身份验证"),
]

for i, (sec, count, purpose) in enumerate(overview, 4):
    for col, val in enumerate([sec, count, purpose], 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.alignment = wrap_align
        c.border = thin_border
        c.font = Font(name="微软雅黑", size=10)

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 40

# -- Sheet 3: 模板格式参考 --
ws3 = wb.create_sheet("模板格式参考")
ws3.merge_cells("A1:B1")
ws3.cell(row=1, column=1, value="问卷摘录 Excel 模板格式说明").font = Font(name="微软雅黑", bold=True, size=13)

template_info = [
    ("列A 序号", "板块字母 + 题号（如B3、C1），方便引用和交叉参照"),
    ("列B 题目/内容", "完整的题目题干，保留换行和强调标记"),
    ("列C 题型", "单选题/多选题/多选题(限N项)/填空/勾选/声明"),
    ("列D 选项", "所有选项逐项列出，A/B/C/D标注"),
    ("列E 备注/来源", "标注信息来源：视频帧号(OCR) 或 官方公告补充"),
    ("颜色规则", "蓝色=板块标题行(白色粗体) / 黄色=声明文本 / 白色=普通题目"),
    ("Sheet2", "问卷结构概览：快速了解各板块题目数量和目的"),
    ("Sheet3", "本说明页"),
]

for i, (key, val) in enumerate(template_info, 3):
    c1 = ws3.cell(row=i, column=1, value=key)
    c1.font = Font(name="微软雅黑", size=10, bold=True)
    c1.alignment = wrap_align
    c1.border = thin_border
    c2 = ws3.cell(row=i, column=2, value=val)
    c2.font = Font(name="微软雅黑", size=10)
    c2.alignment = wrap_align
    c2.border = thin_border

ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 60

# -- Save --
outpath = "D:/claude/异环共存测试招募问卷_完整摘录.xlsx"
wb.save(outpath)
print(f"OK -> {outpath}")
