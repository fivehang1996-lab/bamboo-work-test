import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
info_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
wrap_align = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "绿梦-蚀梦测试招募问卷-TapTap"

ws.merge_cells("A1:E1")
ws.cell(row=1, column=1, value="《绿梦：时空之声》「蚀梦测试」招募问卷 — 完整题目摘录").font = Font(name="微软雅黑", bold=True, size=14)
ws.merge_cells("A2:E2")
ws.cell(row=2, column=1, value="来源：问卷星 https://lvmeng.wjx.cn/vm/OtLskZN.aspx | 类型：招募筛选问卷（TapTap渠道） | 共24题 | 2026年5月").font = Font(name="微软雅黑", size=9, color="666666")

headers = ["序号", "题目/内容", "题型", "选项", "备注"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = wrap_align
    c.border = thin_border

data = []

def S(seq, title):
    data.append((seq, title, "--- 问卷板块 ---", "", ""))

def Q(seq, question, qtype, options, note=""):
    data.append((seq, question, qtype, options, note))

# === 问卷声明 ===
S("A", "问卷声明（首页）")

Q("A1",
  "【招募须知】\n"
  "1. PC/iOS/Android 不计费删档公开测试；iOS招募请前往官网\n"
  "2. 仅针对年满18周岁的成年人开放\n"
  "3. 约需5分钟填写\n"
  "4. 提交前请检查，一经提交无法更改\n"
  "5. 收集个人信息说明（姓名、性别、年龄、职业、联系方式、游戏习惯、设备信息等），严格保密\n"
  "6. 以手机短信/邮件通知资格\n"
  "7. 测试内容为开发中版本，不代表最终品质\n"
  "8. 遵循竹子网络游戏许可及服务协议及隐私政策；联系邮箱: lvmeng_cs@zhuziplay.com",
  "跳转题",
  "开始作答 → 进入问卷\n放弃参与（结束作答）→ 退出",
  "必答")

# === 第一部分：人口学 ===
S("B", "第一部分：基础人口学信息")

Q("B1",
  "请问您的TapTap账号ID是？",
  "文本填空",
  "—",
  "必答。附注：通过站内信通知，一经提交无法修改")

Q("B2",
  "请输入您的出生年月：",
  "日期选择（年月）",
  "—",
  "必答。仅限年满18周岁")

Q("B3",
  "请问您的性别是",
  "单选题",
  "男\n女\n不方便透露",
  "必答")

Q("B4",
  "请问您的职业是",
  "单选题",
  "公司职员/管理者\n退休\n自由职业/个体经营者\n政府事业单位/公务员\n蓝领工人\n学生\n其他\n不方便透露",
  "必答")

Q("B5",
  "请问您的常住城市是",
  "文本填空",
  "—",
  "必答")

# === 第二部分：设备信息 ===
S("C", "第二部分：设备信息")

Q("C1",
  "请问您计划参与测试的设备平台是？",
  "多选题",
  "PC端\n安卓端",
  "必答。附注：可多选，官方视实际情况发放资格")

Q("C2",
  "请问您计划本次参与测试会使用安卓手机设备品牌是？",
  "单选题(+其他填空)",
  "华为\nvivo\nOPPO\n小米\n荣耀\n一加\n红米\niQOO\n三星\n魅族\n其他，请填写：____",
  "必答")

Q("C3",
  "请问您将用于本次测试的手机处理器属于以下哪一个选项？",
  "单选题(+型号填空)",
  "骁龙865/天玑9000/麒麟9000及以下（如：小米10/华为Mate 40 Pro等）\n"
  "骁龙870/888/Gen1/Gen2或天玑9200（如：小米13/红米K60 Pro等）\n"
  "骁龙8Gen3/天玑9300（如红米K80/iQOO Neo10/一加ACE5等）\n"
  "骁龙8Elite/天玑9400及以上（如：小米15/红米K80 Pro/一加Ace5Pro等）\n"
  "不清楚具体配置，请填写手机品牌和型号：____",
  "必答。附注：附查看路径说明")

Q("C4",
  "请问您的手机运行内存(RAM)为？",
  "单选题",
  "4GB及以下\n6GB\n8GB\n12GB\n16GB及以上\n不清楚",
  "必答。附注：附查看路径说明")

Q("C5",
  "请问您能够参与本次测试的PC设备的系统版本是？",
  "单选题",
  "Windows 10 - 32 位\nWindows 10 - 64 位\nWindows 11\n其他 Windows 系统\n其他操作系统（如 Linux、鸿蒙）",
  "必答")

Q("C6",
  "请选择您在测试中将会使用的PC设备的处理器（CPU）型号",
  "下拉选择题",
  "（大列表下拉选择）",
  "必答。附注：附查看路径说明及命名规则说明")

Q("C7",
  "请选择您在测试中将会使用的PC设备的显卡型号",
  "下拉选择题",
  "（大列表下拉选择）",
  "必答。附注：若无对应型号可选性能接近的选项")

# === 第三部分：游戏经验 ===
S("D", "第三部分：游戏经验与品类偏好")

Q("D1",
  "您近半年内主要使用哪些设备玩游戏?",
  "多选题",
  "手机/平板电脑\n笔记本电脑/台式电脑\n家用主机（如Switch、PS、Xbox等）\n以上都没有",
  "必答")

Q("D2",
  "以下手机游戏类型中，您曾深入体验过的有哪些?",
  "多选题(限5项)",
  "Roguelike类（如重生细胞、土豆兄弟等）\n"
  "横版动作类（如火影忍者手游、DNF、归龙潮等）\n"
  "卡牌RPG类（如崩坏：星穹铁道、重返未来：1999、明日方舟、阴阳师等）\n"
  "战棋类（如梦幻模拟战、天地劫：幽城再临等）\n"
  "SLG类（如三国志战略版、三国谋定天下等）\n"
  "回合制MMORPG（如梦幻西游、问道等）\n"
  "动作类（如绝区零、崩坏3、战双帕弥什等）\n"
  "射击类（如三角洲行动、逆战：未来等）\n"
  "沙盒类（如我的世界、迷你世界等）\n"
  "休闲放置类（如剑与远征、三国志幻想大陆等）\n"
  "MOBA类（如王者荣耀、英雄联盟手游等）\n"
  "模拟经营类（如桃园深处有人家、心动小镇等）\n"
  "休闲益智类（如开心消消乐、梦幻家园等）\n"
  "即时制MMORPG（如逆水寒手游、天涯明月刀手游等）\n"
  "自走棋类（如金铲铲之战、多多自走棋等）\n"
  "开放世界类（如原神、鸣潮、幻塔等）\n"
  "其他，请补充：____\n"
  "以上都没有",
  "必答，最多选择5项")

Q("D3",
  "以下手机游戏产品中，您曾深入体验过的有哪些?【横版动作类】",
  "多选题",
  "火影忍者手游\n心渊梦境\n归龙潮\n忍者必须死系列\n重生细胞\n霓虹深渊\n影之刃3\n时空猎人系列\n艾希\n地下城与勇士：起源\n造梦西游系列\n以上都没有",
  "必答")

Q("D4",
  "以下手机游戏产品中，您曾深入体验过的有哪些？【二次元/RPG类】",
  "多选题",
  "白月绮谭\n尘白禁区\n鸣潮\n命运-冠位指定(FGO)\n明日方舟\n崩坏：星穹铁道\n原神\n二重螺旋\n物华弥新\n阴阳师\n明日方舟：终末地\n绝区零\n重返未来：1999\n战双帕弥什\n幻塔\n碧蓝航线\n以上都没有",
  "必答")

Q("D5",
  "以下PC/主机游戏类型中，您最近三个月玩的最多的是？",
  "多选题(限5项)",
  "战争策略类（如帝国时代、魔兽争霸、文明系列等）\n"
  "MOBA（如英雄联盟、DOTA2等）\n"
  "视觉小说/文字冒险类（如命运石之门、隐形守护者、底特律:变人等）\n"
  "集换式卡牌类（如炉石传说、万智牌、游戏王:大师决斗等）\n"
  "平台动作/银河恶魔城类（如奥日、艾希、暗影火炬城等）\n"
  "沙盒建造/生存类（如幻兽帕鲁、我的世界、七日世界、RUST、饥荒等）\n"
  "开放世界类（如塞尔达传说系列、刺客信条系列、GTA5、巫师3等）\n"
  "合作/派对类（如双人成行、糖豆人、猛兽派对等）\n"
  "动作RPG（如只狼、黑暗之魂系列、艾尔登法环、怪物猎人系列等）\n"
  "射击类（如CSGO、无畏契约、守望先锋、Apex、PUBG、三角洲行动、永劫无间等）\n"
  "MMORPG类（如逆水寒、最终幻想14、魔兽世界等）\n"
  "模拟经营类（如沙石镇时光、星露谷物语、露玛岛、双点博物馆等）\n"
  "JRPG类（如碧蓝幻想relink、最终幻想系列等）\n"
  "其他，请补充：____\n"
  "以上都没有",
  "必答，最多选择5项")

Q("D6",
  "以下PC/主机游戏产品中，您曾深度体验过的有哪些?【动作/横版类】",
  "多选题",
  "黑神话悟空\n波斯王子系列\n赤痕：夜之仪式\nDNF端游\n碧蓝幻想：Relink\n只狼：影逝二度\n明末：渊虚之羽\n艾希\n艾尔登法环\n心渊梦境\n苍翼：混沌效应\n空洞骑士：丝之歌\n密特罗德/银河战士系列\n少女与学院城\n空洞骑士\n奥日系列\n黑暗之魂系列\n暗影火炬城\n死亡细胞\n以上都没有",
  "必答")

Q("D7",
  "以下未正式上线的游戏中，有哪些您近期玩过测试版本或非常关注？",
  "多选题(限3项)",
  "伊莫\n粒粒的小人国\n白银之城\n异环\n无限大\n望月\n星布谷地\n蓝色星原：旅谣\n凡应\n归环\n崩坏：因缘精灵\n我最近关注了其他新游戏，请补充：____\n我最近没有关注新游戏",
  "必答，最多选择3项")

# === 第四部分：参与动机 ===
S("E", "第四部分：参与动机与游戏偏好")

Q("E1",
  "请问您主要是被《绿梦：时空之声》的哪些方面吸引，进而参与本次蚀梦测试招募的?",
  "多选题(限3项)",
  "游戏类型（横版动作）\n"
  "剧情世界观\n"
  "战棋玩法\n"
  "朋友、UP主推荐\n"
  "游戏热度高\n"
  "喜欢尝试新游戏\n"
  "动作战斗玩法\n"
  "美术风格\n"
  "有感兴趣的喜欢的角色\n"
  "其他原因，请补充：____",
  "必答，最多选择3项")

Q("E2",
  "请问您在游玩一款新游戏时，哪些游戏乐趣符合您的追求?",
  "多选题(限4项)",
  "【解密】思考成功解决谜题\n"
  "【操作】考验操作精度和速度\n"
  "【策略】需要思考和策略能力\n"
  "【剧情】游戏有精心设计的剧情或故事\n"
  "【人设】角色有吸引人的性格设定、以及配音\n"
  "【画面】精致好看的画面场景\n"
  "【社交】结识新朋友、社交互动\n"
  "【成长】不断提升角色属性和数值，让角色变厉害\n"
  "【收集】收集游戏中的内容（角色/皮肤/道具等）\n"
  "【情感】通过游戏体验/接触到真实的情感关系\n"
  "【沉浸】沉浸在游戏中，忘却现实世界的感觉\n"
  "【探索】探索游戏机制/隐藏内容/彩蛋事件等\n"
  "【休闲】能够让人放松和解压",
  "必答，最多选择4项")

# === 第五部分：联系方式 ===
S("F", "第五部分：联系方式")

Q("F1",
  "请填写您本人的手机号码",
  "文本填空（手机号）",
  "—",
  "必答。附注：用于联系通知测试流程、收集反馈及定位问题")

Q("F2",
  "请填写您参与本次测试将会使用的常用邮箱",
  "文本填空（邮箱）",
  "—",
  "必答。附注：用于后续测试资格信息通知")

# ================================================================
# Write data rows
# ================================================================
for i, (seq, question, qtype, options, note) in enumerate(data, 5):
    row = [seq, question, qtype, options, note]
    for col, val in enumerate(row, 1):
        c = ws.cell(row=i, column=col, value=val)
        c.alignment = wrap_align
        c.border = thin_border
        c.font = Font(name="微软雅黑", size=10)
        if "---" in str(qtype):
            c.fill = section_fill
            c.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        elif "声明" in str(qtype) or "跳转" in str(qtype):
            c.fill = warn_fill

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 65
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 65
ws.column_dimensions["E"].width = 32
ws.freeze_panes = "A5"

# -- Sheet 2: 问卷结构概览 --
ws2 = wb.create_sheet("问卷结构概览")
ws2.merge_cells("A1:C1")
ws2.cell(row=1, column=1, value="《绿梦：时空之声》蚀梦测试招募问卷（TapTap）— 结构概览").font = Font(name="微软雅黑", bold=True, size=13)

for col, h in enumerate(["板块", "题目数", "核心目的"], 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = wrap_align
    c.border = thin_border

overview = [
    ("A. 问卷声明", "1题", "隐私告知与参与确认"),
    ("B. 基础人口学", "5题", "性别/年龄/职业/城市/账号ID"),
    ("C. 设备信息", "7题", "平台/手机品牌/处理器/RAM/PC系统/CPU/显卡"),
    ("D. 游戏经验与偏好", "7题", "手游类型(限5)/横版动作产品/二次元RPG产品/PC主机类型(限5)/动作横版PC产品/关注新游(限3)"),
    ("E. 参与动机", "2题", "游戏吸引因素(限3)+游戏乐趣追求(限4)"),
    ("F. 联系方式", "2题", "手机号 + 邮箱"),
]

for i, (sec, count, purpose) in enumerate(overview, 4):
    for col, val in enumerate([sec, count, purpose], 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.alignment = wrap_align
        c.border = thin_border
        c.font = Font(name="微软雅黑", size=10)

ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 55

# -- Sheet 3: 招募问卷特征分析 --
ws3 = wb.create_sheet("招募问卷设计特征")
ws3.merge_cells("A1:B1")
ws3.cell(row=1, column=1, value="《绿梦》招募问卷 — 设计特征分析").font = Font(name="微软雅黑", bold=True, size=13)

features = [
    ("问卷类型", "测试招募筛选问卷（TapTap渠道版）"),
    ("目标用户", "TapTap平台用户，需年满18周岁"),
    ("问卷长度", "24题，约需5分钟"),
    ("核心筛选维度",
     "① 人口学（年龄/性别/职业/城市）\n"
     "② 设备适配（手机品牌+处理器+RAM / PC系统+CPU+显卡）\n"
     "③ 游戏经验深度（手游类型×手游产品×PC类型×PC产品，四层交叉验证）\n"
     "④ 品类聚焦（横版动作+二次元RPG，与绿梦核心竞品高度相关）\n"
     "⑤ 参与动机（吸引因素+游戏乐趣偏好）"),
    ("竞品矩阵",
     "横版动作类竞品：归龙潮/火影忍者手游/DNF起源/重生细胞/艾希/空洞骑士/死亡细胞/苍翼等\n"
     "二次元RPG竞品：原神/鸣潮/崩铁/绝区零/FGO/明日方舟/重返未来1999/战双等\n"
     "未上线关注：异环/无限大/蓝色星原/望月/白银之城/伊莫等"),
    ("设备分级策略",
     "手机处理器分4档：低(骁龙865及以下)→中低(骁龙870/888)→中高(骁龙8Gen3)→高(骁龙8Elite)\n"
     "PC端采集CPU+显卡具体型号\n"
     "支持PC+安卓双平台（iOS另开渠道），可多选"),
    ("游戏乐趣标签体系",
     "13个乐趣维度：解密/操作/策略/剧情/人设/画面/社交/成长/收集/情感/沉浸/探索/休闲\n"
     "限制4项强制区分优先级"),
    ("与玩法反馈问卷差异",
     "招募问卷：聚焦用户画像+设备+游戏经验+动机，用于资格筛选和分层抽样\n"
     "玩法问卷：聚焦产品内体验（关卡/副本/角色/活动/商业化），用于产品迭代"),
    ("摘录价值",
     "可作为横版动作+二次元RPG品类「招募问卷」的标杆模板，了解网易在TapTap渠道的筛选维度设计和竞品覆盖策略"),
]

for i, (key, val) in enumerate(features, 3):
    c1 = ws3.cell(row=i, column=1, value=key)
    c1.font = Font(name="微软雅黑", size=10, bold=True)
    c1.alignment = wrap_align
    c1.border = thin_border
    c2 = ws3.cell(row=i, column=2, value=val)
    c2.font = Font(name="微软雅黑", size=10)
    c2.alignment = wrap_align
    c2.border = thin_border

ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 68

outpath = "D:/claude/02_问卷摘录/绿梦_蚀梦测试_招募问卷摘录_TapTap.xlsx"
wb.save(outpath)
print(f"OK -> {outpath}")
