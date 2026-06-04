# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = Workbook()
ws = wb.active
ws.title = 'Beta测试体验问卷'

hdr_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
cat_font = Font(name='Microsoft YaHei', size=10, bold=True, color='1F3864')
cat_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
body_font = Font(name='Microsoft YaHei', size=10)
wrap = Alignment(wrap_text=True, vertical='top')
center = Alignment(wrap_text=True, vertical='top', horizontal='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 6
ws.column_dimensions['C'].width = 44
ws.column_dimensions['D'].width = 52
ws.column_dimensions['E'].width = 13
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 30

headers = ['模块', '题号', '题干', '选项', '题型', '必填', '跳转/条件逻辑']
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = center
    c.border = thin_border
ws.row_dimensions[1].height = 24

questions = []

def q(mod, num, stem, opts, qtype, required='是', logic='—'):
    questions.append([mod, num, stem, opts, qtype, required, logic])

# ========== DATA ==========

q('', 'Q1',
  '若您已确认上述内容并无异议，请选择下方"开始作答"并点击【下一页】进行本次问卷作答。',
  '① 开始作答\n② 放弃参与（结束作答）',
  '单选', '是', '选"放弃参与"直接结束问卷')

q('', 'Q2',
  '请问您主要在哪个平台参与本次测试？',
  '① 移动端\n② PC端\n③ PC和移动端均有',
  '单选', '是', '→ 选①隐藏Q3；选②③显示Q3')

q('', 'Q3',
  '请问您在本次PC端测试主要使用的操作设备是？\n（附注：后续问卷填写请以该平台上您的实际体验反馈为主。）',
  '① 键盘和鼠标\n② 手柄\n③ 键鼠及手柄均有使用',
  '单选', '是', '仅在Q2选"PC端"或"双端"时显示')

q('模块1\n整体评价', 'Q4',
  '请问体验到目前为止，您对《绿梦：时空之声》的整体评价如何？',
  '① 远不如我玩的游戏\n② 略差于我玩的游戏\n③ 和我玩的游戏差不多\n④ 略好于我玩的游戏\n⑤ 远好于我玩的游戏',
  '单选（五级）')

q('模块2\n游戏经历\n·手机类型', 'Q5',
  '请问您曾深入体验过以下哪些手机游戏类型？【多选题】',
  '① SLG类（如三国志战略版、三国谋定天下等）\n② 动作类（如绝区零、崩坏3、战双帕弥什等）\n③ 模拟经营类（如桃园深处有人家、心动小镇等）\n④ 回合制MMORPG（如梦幻西游、问道等）\n⑤ 开放世界类（如原神、鸣潮、幻塔等）\n⑥ 战棋类（如梦幻模拟战、天地劫：幽城再临等）\n⑦ 即时制MMORPG（如逆水寒手游、天涯明月刀手游等）\n⑧ 横版动作类（如火影忍者手游、DNF等）\n⑨ 休闲放置类（如剑与远征、三国志幻想大陆等）\n⑩ 沙盒类（如我的世界、迷你世界等）\n⑪ MOBA类（如王者荣耀、英雄联盟手游等）\n⑫ Roguelike类（如重生细胞、土豆兄弟等）\n⑬ 自走棋类（如金铲铲之战、多多自走棋等）\n⑭ 休闲益智类（如开心消消乐、梦幻家园等）\n⑮ 卡牌RPG类（如崩坏：星穹铁道、明日方舟、阴阳师等）\n⑯ 射击类（如三角洲行动、逆战：未来等）\n⑰ 其他，请补充\n⑱ 以上都没有',
  '多选')

q('', 'Q6',
  '请问您曾深入体验过以下哪些手机游戏产品？【多选题】',
  '① 时空猎人系列\n② 苍翼混沌效应\n③ 重生细胞\n④ 影之刃3\n⑤ 造梦西游系列\n⑥ 火影忍者手游\n⑦ 地下城与勇士：起源\n⑧ 归龙潮\n⑨ 心渊梦境\n⑩ 忍者必须死系列\n⑪ 艾希\n⑫ 霓虹深渊\n⑬ 以上都没有',
  '多选')

q('模块2\n游戏经历\n·PC/主机类型', 'Q7',
  '请问您曾深入体验过以下哪些PC/主机游戏类型？【多选题】',
  '① 沙盒建造/生存类（如幻兽帕鲁、我的世界、七日世界、RUST、饥荒等）\n② MMORPG类（如逆水寒、最终幻想14、魔兽世界等）\n③ 射击类（如CSGO、无畏契约、PUBG、三角洲行动等）\n④ 动作RPG（如只狼、黑暗之魂系列、艾尔登法环、怪物猎人系列等）\n⑤ 战争策略类（如帝国时代、魔兽争霸、文明系列等）\n⑥ 集换式卡牌类（如炉石传说、万智牌、游戏王:大师决斗等）\n⑦ 平台动作/银河恶魔城类（如奥日、艾希、暗影火炬城等）\n⑧ 模拟经营类（如沙石镇时光、星露谷物语、露玛岛、双点博物馆等）\n⑨ 合作/派对类（如双人成行、糖豆人、猛兽派对等）\n⑩ MOBA（如英雄联盟、DOTA2等）\n⑪ 视觉小说/文字冒险类（如命运石之门、隐形守护者、底特律:变人等）\n⑫ JRPG类（如碧蓝幻想relink、最终幻想系列等）\n⑬ 开放世界类（如塞尔达传说系列、刺客信条系列、GTA5、巫师3等）\n⑭ 其他，请补充\n⑮ 以上都没有',
  '多选')

q('', 'Q8',
  '请问您曾深入体验过以下哪些PC/主机游戏产品？【多选题】',
  '① 碧蓝幻想：Relink\n② 苍翼：混沌效应\n③ 奥日系列\n④ DNF端游\n⑤ 艾希\n⑥ 空洞骑士\n⑦ 波斯王子系列\n⑧ 黑暗之魂系列\n⑨ 死亡细胞\n⑩ 暗影火炬城\n⑪ 心渊梦境\n⑫ 空洞骑士：丝之歌\n⑬ 只狼：影逝二度\n⑭ 赤痕：夜之仪式\n⑮ 少女与学院城\n⑯ 以上都没有',
  '多选')

q('模块3\n游戏动力', 'Q9',
  '请问当您长期游玩一款游戏时，下列哪些游戏乐趣更符合您的追求？【请选择4项并排序】',
  '① 【剧情】我热衷于挖掘世界观细节、背景故事和角色轶事\n② 【操作】我追求打出华丽连招、精准弹反等操作带来的成就感\n③ 【挑战】我享受挑战高难度BOSS/关卡，追求无伤或S级评价\n④ 【强度】我追求培养出"版本最强"的角色或队伍配置\n⑤ 【策略】我喜欢研究不同敌人的弱点、技能组合和战斗策略\n⑥ 【成就】我喜欢收集成就/藏品图鉴、追求探索度100%\n⑦ 【规划】我喜欢研究并规划长期的养成路径和资源分配\n⑧ 【解谜】我喜欢探索过程中充满挑战性的平台跳跃或机关解谜\n⑨ 【探索】探索地图、寻找宝箱和收集资源是我主要的乐趣之一\n⑩ 【情感】我愿意投入资源培养我喜欢的角色，无论其当前强度如何',
  '排序（选4项）')

q('模块4\n游戏习惯', 'Q10',
  '请问最近三个月，您平均每周花多少时间玩游戏？',
  '① 少于5小时\n② 5~10小时\n③ 11~20小时\n④ 21~30小时\n⑤ 31~40小时\n⑥ 40小时以上',
  '单选')

q('', 'Q11',
  '请问最近三个月，您平均每月在游戏上的消费大约是多少？',
  '① 几乎不花钱\n② 1~50元\n③ 51~200元\n④ 201~500元\n⑤ 501~1000元\n⑥ 1001元~2000元\n⑦ 2001元及以上',
  '单选')

q('模块5\n社媒渠道', 'Q12',
  '请问您通常从哪些渠道了解新游戏的信息？【最多选择4项】',
  '① B站（视频/UP主/直播）\n② 游戏媒体（游民星空/3DM/17173等）\n③ 贴吧 / NGA / 米游社等玩家论坛\n④ TapTap / 好游快爆等游戏社区\n⑤ 微信公众号/朋友圈\n⑥ 应用商店推荐/排行榜\n⑦ 抖音/快手（短视频/直播）\n⑧ 直播平台（斗鱼/虎牙等）\n⑨ 朋友口头推荐\n⑩ 微博\n⑪ 小红书\n⑫ QQ群/微信群\n⑬ 海外平台（YouTube/Twitch/Reddit/Discord）\n⑭ 其他，请补充',
  '多选（限4项）')

q('', 'Q13',
  '请问除微信/QQ外，您日常使用时间最长的社交媒体/内容平台有哪些？【最多选择3项】',
  '① B站\n② 小红书\n③ 快手\n④ 抖音\n⑤ 好游快爆\n⑥ 贴吧\n⑦ 微博\n⑧ 知乎\n⑨ TapTap\n⑩ NGA\n⑪ 其他，请补充',
  '多选（限3项）')

q('', 'Q14',
  '请问您更愿意花时间浏览哪些与游戏相关的内容资讯？【最多选择3项】',
  '① 玩家讨论/社区帖子\n② 图文攻略/专栏文章\n③ 开发者日志/幕后制作\n④ 角色PV/剧情动画\n⑤ 实机游玩视频/攻略\n⑥ 搞笑/二创/同人内容\n⑦ 直播实况\n⑧ 其他，请补充\n⑨ 不怎么看游戏相关内容',
  '多选（限3项）')

q('', 'Q15',
  '请问您更愿意花时间参与哪些与游戏相关的线下活动？【最多选择3项】',
  '① 游戏展会/漫展\n② 线下cosplay聚会\n③ 线下周边售卖\n④ 联名餐饮/快消品打卡\n⑤ 主题快闪店打卡\n⑥ 线下前瞻会/试玩会\n⑦ 其他，请补充\n⑧ 几乎不参与任何线下活动',
  '多选（限3项）')

q('模块6\n人口属性', 'Q16',
  '请问您的性别是？',
  '① 男\n② 女\n③ 不愿透露',
  '单选')

q('', 'Q17',
  '请问您的年龄属于以下哪个区间？',
  '① 18岁以下\n② 18~25岁\n③ 26~35岁\n④ 36~45岁\n⑤ 46~55岁\n⑥ 55岁以上',
  '单选')

q('', 'Q18',
  '请问您的常住地所属哪个城市？',
  '（文本框，自由输入城市名）',
  '填空')

q('模块7\nNPS+开放', 'Q19',
  '请问未来游戏上线后，您向其他朋友推荐《绿梦：时空之声》的意愿有多大？\n（0=不可能，10=极有可能）',
  '0 — 1 — 2 — 3 — 4 — 5 — 6 — 7 — 8 — 9 — 10',
  '量表（0-10）')

q('', 'Q20',
  '非常感谢您完成以上的问卷作答，关于目前的游戏内容和未来的期望，您还有哪些想补充的意见或建议，欢迎在此留下您的宝贵意见。',
  '（多行文本输入框，自由填写）',
  '填空', '否')

# ========== WRITE ==========

# Step 1: Fill all module labels (carry forward)
current_mod = ''
for qi in range(len(questions)):
    if questions[qi][0] and questions[qi][0].strip():
        current_mod = questions[qi][0]
    else:
        questions[qi][0] = current_mod

# Step 2: Clean module labels (remove newlines for single-line display)
# Keep first line only for cleaner display
for qi in range(len(questions)):
    mod = questions[qi][0]
    if '\n' in mod:
        # Keep the module number prefix if it exists
        lines = mod.split('\n')
        questions[qi][0] = lines[0]  # Just first line for column A

row = 2
prev_mod = None
merge_start = None

for qi, qdata in enumerate(questions):
    mod, num, stem, opts, qtype, required, logic = qdata

    vals = [mod, num, stem, opts, qtype, required, logic]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = body_font
        c.alignment = wrap if ci >= 3 else center
        c.border = thin_border

    # Style module column
    ws.cell(row=row, column=1).font = cat_font
    ws.cell(row=row, column=1).fill = cat_fill
    ws.cell(row=row, column=1).alignment = center

    # Row height
    opt_lines = opts.count('\n') + 1
    ws.row_dimensions[row].height = max(50, opt_lines * 15 + 10)

    # Track merge groups
    if mod != prev_mod:
        if merge_start is not None and row - 1 > merge_start:
            ws.merge_cells(start_row=merge_start, start_column=1, end_row=row-1, end_column=1)
        merge_start = row
    prev_mod = mod
    row += 1

# Final merge
if merge_start is not None and row - 1 > merge_start:
    ws.merge_cells(start_row=merge_start, start_column=1, end_row=row-1, end_column=1)

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:G{row-1}'

out = r'D:\claude\绿梦Beta测试体验问卷_题目摘录.xlsx'
wb.save(out)
import os
print(f'Done: {out}')
print(f'Size: {os.path.getsize(out)/1024:.1f} KB')
print(f'Rows: {row-2}')
